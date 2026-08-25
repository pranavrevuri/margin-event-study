#!/usr/bin/env python3
"""
Route 2: parse CME clearing Performance Bond advisories into
data/margin_advisories.csv for the 10 universe products.

Inputs: harvested stub JSONs (scratchpad/advisories/*.json.gz) from the live
notices index. Two advisory eras:
  era A (~2008-2015): full advisory text + rate tables inline in the stub HTML;
  era B (~2016-2020): stub links to a PDF attachment (same URL, .html -> .pdf)
                      whose rows carry Current/New Initial+Maintenance.

Faithful-parse rules: unparseable rows -> data/margin_advisories_exceptions.csv;
no imputation; nothing filtered.
"""
import json
import gzip
import re
import csv
import time
import collections
from pathlib import Path

import pymupdf
from curl_cffi import requests

SP = Path("/private/tmp/claude-501/-Users-nav-Desktop-margin-event-study/f85c660f-ac16-4ed2-9659-3eebf5c848a1/scratchpad")
ADV = SP / "advisories"
PDFD = SP / "adv_pdfs"
PDFD.mkdir(exist_ok=True)
REPO = Path(__file__).resolve().parent.parent
DATA = REPO / "data"

# clearing codes for the universe (advisory tables key rows by CC)
CODES = {"ES": "ES", "EC": "6E", "JY": "6J", "GC": "GC", "SI": "SI",
         "HG": "HG", "CL": "CL", "C": "ZC", "S": "ZS", "21": "ZN"}

RATE_TYPES = {"Spec", "Hedge/Member", "Hedge", "Member"}
CHANGES = {"Increase", "Decrease", "New", "Unchanged", "Change"}
NUM = re.compile(r"^\$?\s*[\d,]+(?:\.\d+)?\s*$")
CUR = re.compile(r"^[A-Z]{3}$")
TIERPAT = re.compile(r"^Mn?ths?\.?\s*\d", re.I)

rows_out = []
exceptions = []
H = {"Referer": "https://www.cmegroup.com/notices.html"}


def num(t):
    return float(t.replace("$", "").replace(",", "").strip())


def parse_dates(text):
    m = re.search(r"Notice Date\s+(\d{1,2}\s+\w+\s+\d{4})", text)
    e = re.search(r"Effective Date\s+(\d{1,2}\s+\w+\s+\d{4})", text)
    def iso(s):
        try:
            return time.strftime("%Y-%m-%d", time.strptime(s, "%d %B %Y"))
        except ValueError:
            return ""
    return iso(m.group(1)) if m else "", iso(e.group(1)) if e else ""


def is_pb(stub):
    hay = (stub.get("title", "") + " " + stub.get("text", "")[:3000]).lower()
    return "performance bond requirement" in hay


def get_pdf(url_html, download=True):
    """download the advisory PDF attachment (cached)"""
    slug = url_html.strip("/").replace("/", "__").replace(".html", ".pdf")
    path = PDFD / slug
    if path.exists() and path.stat().st_size > 0:
        return path
    if not download:
        return None
    pdf_url = "https://www.cmegroup.com" + url_html.replace(".html", ".pdf")
    for a in range(3):
        try:
            r = requests.get(pdf_url, impersonate="chrome", headers=H, timeout=60)
            if r.status_code == 200 and r.content[:5] == b"%PDF-":
                path.write_bytes(r.content)
                return path
            if r.status_code in (403, 429):
                time.sleep(5 * (a + 1)); continue
            return None
        except Exception:
            time.sleep(3 * (a + 1))
    return None


def parse_pdf_rows(path, notice_date, effective_date, src):
    """era-B PDF: rows [RateType, Change, CUR, 4 numbers, (tier), CC], each row
    bound to its nearest preceding product header. Outright headers end with the
    clearing code '(CC)'; spread/combination headers (e.g. '... Month 1 vs 2
    (E-MINI S&P 500 FUTURES)') do not, and their rows are logged, not kept."""
    doc = pymupdf.open(str(path))
    toks = []
    has_outright = False
    for p in doc:
        t = p.get_text()
        if "Outright Rates" in t and "Inter-commodity" not in t.split("Outright Rates")[0][-40:]:
            has_outright = has_outright or ("- Outright Rates" in t or "Outright Rates \n" in t)
        toks += [l.strip() for l in t.split("\n") if l.strip()]
    doc.close()
    n = len(toks)
    i = 0
    found = 0
    last_header = ""
    while i < n:
        t = toks[i]
        if t not in RATE_TYPES and "(" in t and ")" in t and not NUM.match(t):
            last_header = t
            i += 1
            continue
        if t in RATE_TYPES:
            j = i + 1
            if j < n and toks[j] in CHANGES:
                change = toks[j]; j += 1
                if j < n and CUR.match(toks[j]):
                    cur = toks[j]; j += 1
                    nums = []
                    while j < n and len(nums) < 4 and NUM.match(toks[j]):
                        nums.append(num(toks[j])); j += 1
                    tier = ""
                    if j < n and not re.match(r"^[A-Z0-9]{1,4}$", toks[j]):
                        tier = toks[j].strip(); j += 1   # e.g. 'Mnth 1', 'All Months', 'Days 7-15'
                    cc = toks[j] if j < n and re.match(r"^[A-Z0-9]{1,4}$", toks[j]) else None
                    if len(nums) == 4 and cc is not None:
                        if cc in CODES:
                            outright = last_header.rstrip().endswith(f"({cc})")
                            if outright:
                                rows_out.append(dict(
                                    product=CODES[cc], clearing_code=cc,
                                    notice_date=notice_date, effective_date=effective_date,
                                    rate_type=t, change=change, currency=cur, tier=tier,
                                    current_initial=nums[0], current_maintenance=nums[1],
                                    new_initial=nums[2], new_maintenance=nums[3],
                                    source=src, era="B-pdf"))
                            else:
                                exceptions.append((src, f"{last_header} | {t} {change} {nums} {cc}",
                                                   "non-outright (spread/combination) row for universe code; excluded from outright CSV"))
                        found += 1
                        i = j + 1
                        continue
                    exceptions.append((src, " | ".join(toks[i:i + 9]), "malformed PDF rate row"))
            i += 1
        else:
            i += 1
    if found == 0 and not has_outright:
        return -1   # advisory contains no outright-rate section (spread/credit-only): handled
    return found


def parse_inline_text(stub, notice_date, effective_date, src):
    """era-A: flattened text rows like
    'Spec Increase $4,320 $3,200 $5,400 $4,000' under a 'Product (CC)' heading."""
    text = stub.get("text", "")
    found = 0
    # product sections: "Name (CC)" ... rate rows until next section
    matches = list(re.finditer(r"([A-Z][A-Za-z0-9 .,&/\-']{2,60})\s*\(([A-Z0-9]{1,4})\)", text))
    for k, m in enumerate(matches):
        cc = m.group(2)
        if cc not in CODES:
            continue
        seg = text[m.end(): matches[k + 1].start() if k + 1 < len(matches) else m.end() + 2500]
        for rm in re.finditer(
                r"(Spec|Hedge/Member|Hedge|Member)\s+(Increase|Decrease|New|Unchanged)\s+"
                r"\$?([\d,]+(?:\.\d+)?)\s+\$?([\d,]+(?:\.\d+)?)\s+\$?([\d,]+(?:\.\d+)?)\s+\$?([\d,]+(?:\.\d+)?)", seg):
            rows_out.append(dict(
                product=CODES[cc], clearing_code=cc,
                notice_date=notice_date, effective_date=effective_date,
                rate_type=rm.group(1), change=rm.group(2), currency="USD", tier="",
                current_initial=num(rm.group(3)), current_maintenance=num(rm.group(4)),
                new_initial=num(rm.group(5)), new_maintenance=num(rm.group(6)),
                source=src, era="A-inline"))
            found += 1
    return found



def find_attachment(stub):
    """cached attachment path for a stub: pdf (same slug) or the stub's linked xlsx/pdf"""
    base_pdf = PDFD / (stub["url"].strip("/").replace("/", "__").replace(".html", ".pdf"))
    if base_pdf.exists() and base_pdf.stat().st_size > 0:
        return base_pdf
    base = stub["url"].rsplit("/", 1)[-1].replace(".html", "").lower()
    for h, a in stub.get("links", []):
        hl = h.lower()
        if base in hl and hl.endswith((".pdf", ".xlsx", ".xls")):
            pth = PDFD / h.strip("/").replace("/", "__")
            if pth.exists() and pth.stat().st_size > 0:
                return pth
    return None


def parse_xlsx_rows(path, notice_date, effective_date, src):
    """2016-era per-product-family XLSX: TOC (product codes, scaling factors,
    effective date) + 'Outright' sheet (per contract-month Old/New maintenance).
    TOC states: rates are maintenance; initial = 110% (not imputed here)."""
    import openpyxl
    from datetime import datetime, date as _date
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        exceptions.append((src, "", f"unreadable xlsx: {e}"))
        return 0
    names = wb.sheetnames
    toc_name = next((s for s in names if "content" in s.lower()), names[0])
    out_name = next((s for s in names if s.lower().strip() == "outright"), None)
    if out_name is None:
        exceptions.append((src, str(names), "xlsx has no Outright sheet (spread-only advisory); handled"))
        wb.close()
        return -1
    eff = effective_date
    prods = {}  # combined commodity -> list of (product_code, scale)
    header_seen = False
    for row in wb[toc_name].iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        for c in cells:
            if isinstance(c, str) and "Effective Date:" in c:
                m = re.search(r"Effective Date:\s*(\d{1,2})/(\d{1,2})/(\d{4})", c)
                if m:
                    eff = f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        vals = list(row)
        if any(isinstance(c, str) and c.strip() == "Product Code" for c in vals):
            header_seen = True
            continue
        if header_seen and len(vals) >= 5 and vals[1] and vals[2]:
            comb, pcode, name, scale = vals[1], vals[2], vals[3], vals[4]
            if isinstance(pcode, str) and isinstance(comb, str) and isinstance(scale, (int, float)):
                prods.setdefault(comb.strip(), []).append((pcode.strip(), float(scale)))
    found = 0
    for row in wb[out_name].iter_rows(values_only=True):
        vals = [c for c in row if c is not None]
        if len(vals) < 5 or not isinstance(vals[0], str):
            continue
        comb = vals[0].strip()
        try:
            period = str(vals[1])
            pcode_month = vals[2]
            old_m = float(vals[3])
            new_m = float(vals[4])
        except (ValueError, TypeError):
            continue
        found += 1   # structurally-valid outright row (any product)
        if comb not in prods:
            continue
        month = pcode_month.date().isoformat() if isinstance(pcode_month, datetime) else str(pcode_month)[:10]
        for pcode, scale in prods[comb]:
            if pcode not in CODES:
                continue
            rows_out.append(dict(
                product=CODES[pcode], clearing_code=pcode,
                notice_date=notice_date, effective_date=eff,
                rate_type="Maintenance", change=("Increase" if new_m > old_m else ("Decrease" if new_m < old_m else "Unchanged")),
                currency="USD", tier=f"Period {period} ({month})",
                current_initial="", current_maintenance=old_m * scale,
                new_initial="", new_maintenance=new_m * scale,
                source=src, era="B-xlsx"))
    wb.close()
    return found


def main():
    from concurrent.futures import ThreadPoolExecutor

    stubs = sorted(ADV.glob("*.json.gz"))
    print(f"{len(stubs)} stubs harvested", flush=True)
    n_pb = n_pdf_ok = n_inline = 0
    per_year = collections.Counter()
    pb_stubs = []
    for sp_ in stubs:
        stub = json.load(gzip.open(sp_, "rt"))
        if is_pb(stub):
            pb_stubs.append(stub)
    print(f"PB stubs: {len(pb_stubs)}", flush=True)

    # split by era, prefetch era-B PDFs in parallel
    needs_pdf = [s for s in pb_stubs
                 if not re.search(r"(Spec|Hedge/Member)\s+(Increase|Decrease)\s+\$?[\d,]+", s.get("text", ""))]
    print(f"era-B (PDF attachment) stubs to fetch: {len(needs_pdf)}", flush=True)
    import os
    if os.environ.get("ADV_CACHE_ONLY") != "1":
        with ThreadPoolExecutor(max_workers=6) as ex:
            for i, _ in enumerate(ex.map(lambda s: get_pdf(s["url"]), needs_pdf), 1):
                if i % 50 == 0:
                    print(f"pdf prefetch {i}/{len(needs_pdf)}", flush=True)

    per_year_ok = collections.Counter()
    for stub in pb_stubs:
        n_pb += 1
        nd, ed = parse_dates(stub.get("text", ""))
        src = stub["url"]
        yr = (ed or nd or "????")[:4]
        per_year[yr] += 1
        got = parse_inline_text(stub, nd, ed, src)
        if got > 0:
            n_inline += 1
            per_year_ok[yr] += 1
            continue
        # era B: attachment (cache-only here; downloads happen in the prefetch stage)
        att = find_attachment(stub)
        if att is None:
            exceptions.append((src, "", "PB advisory attachment not yet fetched (rate-limited); rerun to retry"))
            continue
        if str(att).lower().endswith((".xlsx", ".xls")):
            got = parse_xlsx_rows(att, nd, ed, src)
        else:
            got = parse_pdf_rows(att, nd, ed, src)
        if got > 0 or got == -1:
            n_pdf_ok += 1
            per_year_ok[yr] += 1
            if got == -1:
                exceptions.append((src, "", "no outright-rate content (spread/credit-only advisory); counted handled"))
        else:
            exceptions.append((src, "", "PB advisory attachment yielded no rate rows (format not recognized)"))

    cols = ["product", "clearing_code", "notice_date", "effective_date", "rate_type",
            "change", "currency", "tier", "current_initial", "current_maintenance",
            "new_initial", "new_maintenance", "source", "era"]
    with open(DATA / "margin_advisories.csv", "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerows(sorted(rows_out, key=lambda r: (r["product"], r["effective_date"], r["tier"], r["rate_type"])))
    with open(DATA / "margin_advisories_exceptions.csv", "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["source", "raw", "reason"])
        w.writerows(exceptions)

    with open(DATA / "margin_advisories_completeness.csv", "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["year", "pb_advisories", "parsed_ok", "parse_rate"])
        for yr in sorted(per_year):
            w.writerow([yr, per_year[yr], per_year_ok.get(yr, 0),
                        round(per_year_ok.get(yr, 0) / per_year[yr], 3)])
    print(f"PB advisories: {n_pb} (inline-parsed {n_inline}, pdf-parsed {n_pdf_ok})")
    print("PB advisories by effective year:", dict(sorted(per_year.items())))
    print(f"universe rate rows: {len(rows_out)}; exceptions: {len(exceptions)}")
    per_prod = collections.Counter(r["product"] for r in rows_out)
    print("rows by product:", dict(sorted(per_prod.items())))


if __name__ == "__main__":
    main()
